from dataclasses import dataclass

import openpyxl
from django.db.models import Count, Max, Prefetch
from django.http import HttpResponse
from django.utils import timezone

from database.models import DbConfigSource, DbConfigSourceFieldDefinition, DbConfigSourceFieldValue
from facts.dynamic_fields import coerce_fact_value

DB_UNIQUE_NAME_HEADER = "db_unique_name"
MAX_ROWS = 2000

# dashboard/excel_import.py(자산)의 FIXED_REFERENCE_LABELS와 동일한 취지 - db_unique_name처럼
# 매칭 키는 아니지만 DbConfigSourceFieldDefinition이 아니라 대시보드 화면 전용 하드코딩
# 컬럼이라 동적 필드 목록엔 안 걸리는 참고용 컬럼. AUTO 필드와 같은 취지로 다운로드엔 싣되
# 업로드 시엔 항상 무시(편집 대상 아님, 값 자체가 push/재계산으로만 바뀜) - 헤더 매칭은
# 위치가 아니라 이름으로만 하므로 종류~인스턴스 수(목록 앞쪽)와 최근 변경일/최근 반영일
# (목록 뒤쪽)을 한 목록에 같이 둬도 무방하다.
LEADING_REFERENCE_LABELS = ["종류", "DB명", "Role", "Open Mode"]
TRAILING_REFERENCE_LABELS = ["인스턴스 수", "최근 변경일", "최근 반영일"]
FIXED_REFERENCE_LABELS = LEADING_REFERENCE_LABELS + TRAILING_REFERENCE_LABELS


class ImportFileError(Exception):
    """헤더가 잘못됐거나 컬럼을 매칭할 수 없는 등, 파일 자체를 거부해야 하는 경우."""


@dataclass
class PendingUpdate:
    row_number: int
    source_id: int
    db_unique_name: str
    field_id: int
    field_label: str
    old_value: str
    new_value: str


@dataclass
class InvalidCell:
    row_number: int
    db_unique_name: str
    field_label: str
    raw_value: str


@dataclass
class ImportResult:
    updates: list[PendingUpdate]
    unmatched_names: list[tuple[int, str]]
    invalid_cells: list[InvalidCell]


def _dynamic_fields() -> list[DbConfigSourceFieldDefinition]:
    """전체 동적 필드(AUTO+MANUAL) - 다운로드에는 둘 다 참고용으로 실어서 대시보드 화면과
    같은 값을 보여주고, 실제 반영 대상은 이 중 MANUAL만이다(아래 parse_db_config_workbook)."""
    return list(
        DbConfigSourceFieldDefinition.objects.filter(is_visible=True).order_by("sort_order", "id")
    )


def _dynamic_fields_by_label(
    dynamic_fields: list[DbConfigSourceFieldDefinition],
) -> dict[str, DbConfigSourceFieldDefinition]:
    labels = [fd.label for fd in dynamic_fields]
    duplicate_labels = sorted({label for label in labels if labels.count(label) > 1})
    if duplicate_labels:
        raise ImportFileError(
            "다음 라벨을 쓰는 필드가 여러 개 등록돼 있어 엑셀 헤더와 매칭할 수 없습니다: "
            f"{', '.join(duplicate_labels)}. admin에서 라벨을 정리한 뒤 다시 시도해주세요."
        )
    return {fd.label: fd for fd in dynamic_fields}


def _current_stored_value(source, field_definition: DbConfigSourceFieldDefinition):
    return next(
        (v for v in source.field_values.all() if v.field_definition_id == field_definition.id), None
    )


def _current_value_display(source, field_definition: DbConfigSourceFieldDefinition) -> str:
    value = _current_stored_value(source, field_definition)
    if value is None:
        return ""
    for candidate in (value.value_text, value.value_number, value.value_date):
        if candidate not in (None, ""):
            return str(candidate)
    return ""


def _cell_datetime(value):
    """tz-aware datetime은 openpyxl이 그대로 저장 못 한다(Excel엔 타임존 개념이 없어 저장
    자체가 거부됨) - 대시보드 화면과 같은 로컬 시각으로 보이도록 변환 후 tzinfo를 제거한다."""
    if value is None:
        return ""
    return timezone.localtime(value).replace(tzinfo=None)


def export_db_config_workbook() -> HttpResponse:
    """전체 DB의 db_unique_name + 동적 필드(AUTO+MANUAL) 현재 값을 xlsx로 내려준다(자산/
    시스템 목록과 동일한 취지 - 몇 칸만 고쳐서 재업로드하는 흐름). AUTO 필드는 대시보드
    화면과 맥락을 맞춰 참고용으로 같이 싣지만 편집 대상이 아니라서, 값을 고쳐서 올려도
    parse_db_config_workbook이 조용히 무시한다(반영은 MANUAL 필드만)."""
    dynamic_fields = _dynamic_fields()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "DB"
    sheet.append(
        [
            DB_UNIQUE_NAME_HEADER,
            *LEADING_REFERENCE_LABELS,
            *[fd.label for fd in dynamic_fields],
            *TRAILING_REFERENCE_LABELS,
        ]
    )

    sources = (
        DbConfigSource.objects.prefetch_related(
            Prefetch(
                "field_values", queryset=DbConfigSourceFieldValue.objects.select_related("field_definition")
            )
        )
        .annotate(
            instance_count=Count("instances", distinct=True),
            last_changed_at=Max("revisions__detected_at"),
        )
        .order_by("db_unique_name")
    )
    for source in sources:
        sheet.append([
            source.db_unique_name,
            source.get_kind_display(),
            source.db_name,
            source.database_role,
            source.open_mode,
            *[_current_value_display(source, fd) for fd in dynamic_fields],
            source.instance_count,
            _cell_datetime(source.last_changed_at),
            _cell_datetime(source.last_pushed_at),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cmdb_database.xlsx"'
    workbook.save(response)
    return response


def parse_db_config_workbook(uploaded_file) -> ImportResult:
    try:
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True)
    except Exception as exc:  # openpyxl은 손상 파일에 다양한 예외를 던짐
        raise ImportFileError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc

    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)

    header = next(rows_iter, None)
    if not header or str(header[0] or "").strip() != DB_UNIQUE_NAME_HEADER:
        raise ImportFileError(f"첫 번째 컬럼 헤더는 '{DB_UNIQUE_NAME_HEADER}'이어야 합니다.")

    # AUTO 필드 헤더도 알려진 컬럼으로 인정한다(다운로드가 AUTO도 참고용으로 같이 싣기
    # 때문에, 안 고치고 그대로 재업로드해도 "인식 못하는 컬럼"으로 걸리면 안 됨) - 실제
    # 반영 여부는 아래 행 처리 루프에서 source로 다시 가른다.
    fields_by_label = _dynamic_fields_by_label(_dynamic_fields())

    column_fields: list[DbConfigSourceFieldDefinition | None] = []
    unknown_headers = []
    for col_header in header[1:]:
        col_header = str(col_header).strip() if col_header is not None else ""
        if not col_header or col_header in FIXED_REFERENCE_LABELS:
            column_fields.append(None)
            continue
        field_definition = fields_by_label.get(col_header)
        if field_definition is None:
            unknown_headers.append(col_header)
        column_fields.append(field_definition)

    if unknown_headers:
        raise ImportFileError(
            "다음 컬럼은 등록된 필드와 매칭되지 않습니다: " + ", ".join(unknown_headers)
        )

    choice_values_by_field_id = {
        fd.id: set(fd.choices.values_list("value", flat=True))
        for fd in column_fields
        if fd is not None and fd.value_type == DbConfigSourceFieldDefinition.ValueType.CHOICE
    }

    raw_rows = []
    names = set()
    for row_number, row in enumerate(rows_iter, start=2):
        if row is None or all(cell in (None, "") for cell in row):
            continue
        db_unique_name = str(row[0]).strip() if row[0] is not None else ""
        if not db_unique_name:
            continue
        names.add(db_unique_name)
        raw_rows.append((row_number, db_unique_name, row[1:]))

        if len(raw_rows) > MAX_ROWS:
            raise ImportFileError(f"한 번에 업로드 가능한 최대 행 수({MAX_ROWS}건)를 초과했습니다.")

    sources_by_name = {
        source.db_unique_name: source
        for source in DbConfigSource.objects.filter(db_unique_name__in=names).prefetch_related(
            Prefetch(
                "field_values", queryset=DbConfigSourceFieldValue.objects.select_related("field_definition")
            )
        )
    }

    updates: list[PendingUpdate] = []
    unmatched_names: list[tuple[int, str]] = []
    invalid_cells: list[InvalidCell] = []

    for row_number, db_unique_name, values in raw_rows:
        source = sources_by_name.get(db_unique_name)
        if source is None:
            unmatched_names.append((row_number, db_unique_name))
            continue

        for field_definition, cell_value in zip(column_fields, values):
            if (
                field_definition is None
                or field_definition.source != DbConfigSourceFieldDefinition.Source.MANUAL
                or cell_value in (None, "")
            ):
                # AUTO 필드는 참고용으로만 실려있는 컬럼이라 값이 있어도 반영하지 않는다 -
                # 어차피 다음 push 때 조용히 덮어써지므로.
                continue

            defaults = coerce_fact_value(cell_value, field_definition.value_type)
            parse_failed = (
                field_definition.value_type
                in (DbConfigSourceFieldDefinition.ValueType.NUMBER, DbConfigSourceFieldDefinition.ValueType.DATE)
                and all(v is None for v in defaults.values())
            )
            invalid_choice = field_definition.value_type == DbConfigSourceFieldDefinition.ValueType.CHOICE and (
                str(cell_value) not in choice_values_by_field_id.get(field_definition.id, set())
            )
            if parse_failed or invalid_choice:
                invalid_cells.append(
                    InvalidCell(
                        row_number=row_number,
                        db_unique_name=db_unique_name,
                        field_label=field_definition.label,
                        raw_value=str(cell_value),
                    )
                )
                continue

            stored_value = _current_stored_value(source, field_definition)
            unchanged = stored_value is not None and all(
                getattr(stored_value, key) == value for key, value in defaults.items()
            )
            if unchanged:
                continue

            updates.append(
                PendingUpdate(
                    row_number=row_number,
                    source_id=source.id,
                    db_unique_name=db_unique_name,
                    field_id=field_definition.id,
                    field_label=field_definition.label,
                    old_value=_current_value_display(source, field_definition),
                    new_value=str(cell_value),
                )
            )

    return ImportResult(
        updates=updates, unmatched_names=unmatched_names, invalid_cells=invalid_cells
    )


def apply_updates(payload: list[dict]) -> tuple[int, int]:
    """미리보기에서 확정된 항목을 실제로 반영. (반영된 값 개수, 영향받은 DB 수)를 반환."""
    field_definitions = {
        fd.id: fd
        for fd in DbConfigSourceFieldDefinition.objects.filter(
            source=DbConfigSourceFieldDefinition.Source.MANUAL, is_visible=True
        )
    }
    source_ids = {item["source_id"] for item in payload}
    sources = {source.id: source for source in DbConfigSource.objects.filter(id__in=source_ids)}

    changed_source_ids = set()
    applied = 0
    for item in payload:
        field_definition = field_definitions.get(item["field_id"])
        source = sources.get(item["source_id"])
        if field_definition is None or source is None:
            continue

        defaults = coerce_fact_value(item["new_value"], field_definition.value_type)
        DbConfigSourceFieldValue.objects.update_or_create(
            source=source, field_definition=field_definition, defaults=defaults
        )
        changed_source_ids.add(source.id)
        applied += 1

    return applied, len(changed_source_ids)
