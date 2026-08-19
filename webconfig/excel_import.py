from dataclasses import dataclass

import openpyxl
from django.http import HttpResponse

from core.models import Service
from core.reconciliation import normalize_hostname
from webconfig.models import ApacheVhost, NginxVhost, WebConfigSource, WebServiceDomain, WebtobVhost

# kind별 vhost 모델 - service_name은 vhost 쪽이 원본이고 WebServiceDomain은 복사본이라
# 엑셀 반영 시 둘 다 갱신해야 한다(webconfig/sync.py의 sync_service_domains와 같은 이유).
VHOST_MODELS = {
    WebConfigSource.Kind.WEBTOB: WebtobVhost,
    WebConfigSource.Kind.APACHE: ApacheVhost,
    WebConfigSource.Kind.NGINX: NginxVhost,
}

HOSTNAME_HEADER = "hostname"
VHOST_HEADER = "vhost"
DOMAIN_HEADER = "domain"
ALIASES_HEADER = "aliases"
PORT_HEADER = "포트"
SERVICE_NAME_HEADER = "서비스명"
KIND_HEADER = "솔루션"
VERSION_HEADER = "솔루션버전"
FIX_HEADER = "Fix"
MAX_ROWS = 2000

# 다운로드 컬럼 순서를 서비스 조회 화면(webservice_list.html의 WEB 표: 서비스명/도메인/
# 포트/Hostname/솔루션)과 맞추기 위해, 헤더는 위치가 아니라 이름으로 찾는다(hostname/vhost도
# 예외 없이 동일 방식) - 컬럼 순서를 화면과 다르게 재배치해도 매칭 로직이 깨지지 않는다.
# 도메인은 vhost 하나당 유일하지 않을 수 있어서(같은 도메인을 http/https vhost 두 개가
# 나눠 쓰는 경우가 실제로 있음, 예: vhost1/vhost1_ssl) 매칭 키로 못 쓴다. domain/aliases/
# 포트/솔루션(kind)/솔루션버전/Fix는 사람이 알아보기 위한 참고용 컬럼으로만 두고, 실제
# 매칭은 (hostname, vhost) 조합으로 한다 - vhost 이름은 source당 유일(WebtobVhost의 unique
# constraint)이라 안전하다. vhost 자체는 화면엔 안 보이지만(도메인은 vhost 두 개가 나눠
# 쓸 수 있어 화면 표시엔 불필요) 매칭 키라 엑셀엔 여전히 필요 - Hostname 바로 뒤에 둔다.
# 솔루션(kind)/솔루션버전/Fix는 이제 push에서 자동 추출되는 AUTO 값이라(webconfig/kind는
# 설정 파일 종류, version_extract.py) 엑셀로 수정해도 다음 push 때 덮어써지므로, 여기서도
# domain/aliases와 같은 참고용 컬럼으로만 두고 업로드 시엔 무시한다.
_HEADER_ROLES = {
    HOSTNAME_HEADER: "hostname",
    VHOST_HEADER: "vhost",
    DOMAIN_HEADER: "domain",
    ALIASES_HEADER: "aliases",
    PORT_HEADER: "port",
    SERVICE_NAME_HEADER: "service_name",
    KIND_HEADER: "kind",
    VERSION_HEADER: "version",
    FIX_HEADER: "fix",
}


class ImportFileError(Exception):
    """헤더가 잘못됐거나 컬럼을 매칭할 수 없는 등, 파일 자체를 거부해야 하는 경우."""


@dataclass
class ServiceNameUpdate:
    row_number: int
    service_domain_id: int
    hostname: str
    vhost_name: str
    old_value: str
    new_value: str


@dataclass
class UnmatchedRow:
    row_number: int
    hostname: str
    vhost_name: str


@dataclass
class UnknownServiceRow:
    """(hostname, vhost)는 매칭됐지만 새로 지정한 서비스명이 등록된 core.Service에 없는 경우
    - 대시보드 인라인 편집(_resolve_service)과 동일하게, 오타로 새 Service가 조용히 생기는
    걸 막기 위해 여기서도 get_or_create를 안 쓴다. 확정 대상에서 빼고 미리보기에 따로 보여준다."""

    row_number: int
    hostname: str
    vhost_name: str
    attempted_name: str


@dataclass
class ServiceImportResult:
    service_name_updates: list[ServiceNameUpdate]
    unmatched_rows: list[UnmatchedRow]
    unknown_service_rows: list[UnknownServiceRow]


def export_service_workbook() -> HttpResponse:
    """지금 서비스 조회 화면의 전체 데이터를 그대로 xlsx로 내려준다(검색 필터 무시 - 완전한
    사본을 줘야 재업로드할 때 행이 빠져서 헷갈리는 일이 없음). 컬럼 순서는 화면(WEB 표: 서비스명/
    도메인/포트/Hostname/솔루션)과 맞추고, 화면엔 없지만 매칭 키/참고용으로 필요한 vhost/
    aliases/솔루션버전/Fix는 각자와 가장 관련 있는 컬럼 바로 뒤에 붙인다. 헤더가 그대로 업로드
    형식이라(이름으로 컬럼을 찾으므로 순서 자체는 문제되지 않음) 다운로드 → 몇 칸만 고쳐서
    재업로드하는 흐름으로 쓰라고 만든 것."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "서비스"
    sheet.append(
        [
            SERVICE_NAME_HEADER,
            DOMAIN_HEADER,
            ALIASES_HEADER,
            PORT_HEADER,
            HOSTNAME_HEADER,
            VHOST_HEADER,
            KIND_HEADER,
            VERSION_HEADER,
            FIX_HEADER,
        ]
    )

    rows = WebServiceDomain.objects.select_related("source", "source__asset").order_by(
        "source__asset__hostname", "domain"
    )
    for row in rows:
        sheet.append(
            [
                row.service_name,
                row.domain,
                row.aliases,
                row.port,
                row.source.asset.hostname,
                row.vhost_name,
                row.source.get_kind_display(),
                row.source.solution_version,
                row.source.solution_fix,
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cmdb_services.xlsx"'
    workbook.save(response)
    return response


def parse_service_workbook(uploaded_file) -> ServiceImportResult:
    try:
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True)
    except Exception as exc:  # openpyxl은 손상 파일에 다양한 예외를 던짐
        raise ImportFileError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc

    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)

    header = next(rows_iter, None)
    if not header:
        raise ImportFileError("헤더 행이 없습니다.")

    # 컬럼을 위치가 아니라 이름으로 찾는다 - 다운로드 컬럼 순서를 화면에 맞춰 바꿔도(위
    # export_service_workbook 참고) hostname/vhost가 몇 번째 컬럼이든 상관없이 매칭된다.
    column_roles: list[str | None] = []
    unknown_headers = []
    for col_header in header:
        col_header = str(col_header).strip() if col_header is not None else ""
        if not col_header:
            column_roles.append(None)
            continue
        role = _HEADER_ROLES.get(col_header)
        if role is None:
            unknown_headers.append(col_header)
        column_roles.append(role)

    if unknown_headers:
        raise ImportFileError("다음 컬럼을 인식할 수 없습니다: " + ", ".join(unknown_headers))

    if "hostname" not in column_roles or "vhost" not in column_roles:
        raise ImportFileError(
            f"'{HOSTNAME_HEADER}', '{VHOST_HEADER}' 컬럼이 모두 있어야 합니다."
        )

    hostname_col = column_roles.index("hostname")
    vhost_col = column_roles.index("vhost")
    service_name_col = column_roles.index("service_name") if "service_name" in column_roles else None

    raw_rows = []
    for row_number, row in enumerate(rows_iter, start=2):
        if row is None or all(cell in (None, "") for cell in row):
            continue
        hostname = normalize_hostname(str(row[hostname_col])) if row[hostname_col] is not None else ""
        vhost_name = (
            str(row[vhost_col]).strip() if len(row) > vhost_col and row[vhost_col] is not None else ""
        )
        if not hostname:
            continue
        raw_rows.append((row_number, hostname, vhost_name, row))

        if len(raw_rows) > MAX_ROWS:
            raise ImportFileError(f"한 번에 업로드 가능한 최대 행 수({MAX_ROWS}건)를 초과했습니다.")

    hostnames = {r[1] for r in raw_rows}
    service_domains = list(
        WebServiceDomain.objects.filter(source__asset__hostname__in=hostnames).select_related(
            "source", "source__asset"
        )
    )
    row_map = {(sd.source.asset.hostname, sd.vhost_name): sd for sd in service_domains}
    # 대시보드 인라인 편집(_resolve_service)과 동일한 정책 - 오타로 새 Service가 조용히
    # 생기는 걸 막기 위해 기존에 등록된 서비스명만 반영 대상으로 허용한다(새 서비스는
    # "서비스" 탭의 "새 서비스 등록"으로 먼저 만들어야 함).
    existing_service_names = set(Service.objects.values_list("name", flat=True))

    service_name_updates: list[ServiceNameUpdate] = []
    unmatched_rows: list[UnmatchedRow] = []
    unknown_service_rows: list[UnknownServiceRow] = []

    for row_number, hostname, vhost_name, row in raw_rows:
        service_domain = row_map.get((hostname, vhost_name))
        if service_domain is None:
            unmatched_rows.append(
                UnmatchedRow(row_number=row_number, hostname=hostname, vhost_name=vhost_name)
            )
            continue

        if service_name_col is not None and service_name_col < len(row):
            cell_value = row[service_name_col]
            if cell_value not in (None, ""):
                new_value = str(cell_value).strip()
                if new_value != service_domain.service_name:
                    if new_value not in existing_service_names:
                        unknown_service_rows.append(
                            UnknownServiceRow(
                                row_number=row_number,
                                hostname=hostname,
                                vhost_name=vhost_name,
                                attempted_name=new_value,
                            )
                        )
                        continue
                    service_name_updates.append(
                        ServiceNameUpdate(
                            row_number=row_number,
                            service_domain_id=service_domain.id,
                            hostname=hostname,
                            vhost_name=vhost_name,
                            old_value=service_domain.service_name,
                            new_value=new_value,
                        )
                    )

    return ServiceImportResult(
        service_name_updates=service_name_updates,
        unmatched_rows=unmatched_rows,
        unknown_service_rows=unknown_service_rows,
    )


def apply_service_updates(payload: list[dict]) -> tuple[int, int]:
    """미리보기에서 확정된 항목을 실제로 반영. (반영된 값 개수, 영향받은 자산 수)를 반환."""
    service_name_items = [p for p in payload if p.get("kind") == "service_name"]

    changed_asset_ids = set()
    applied = 0

    if service_name_items:
        service_domains = {
            sd.id: sd
            for sd in WebServiceDomain.objects.filter(
                id__in=[item["service_domain_id"] for item in service_name_items]
            ).select_related("source")
        }
        for item in service_name_items:
            service_domain = service_domains.get(item["service_domain_id"])
            if service_domain is None:
                continue
            new_value = item["new_value"]
            service = None
            if new_value:
                # 미리보기(parse_service_workbook)가 이미 등록된 서비스명만 여기까지 넘겨주므로
                # 정상적으로는 항상 찾아진다 - 그 사이 삭제된 경우를 대비해 방어적으로 None 허용
                # (get_or_create로 새로 만들지는 않음 - 오타로 새 Service가 생기는 걸 막는 게 목적).
                service = Service.objects.filter(name=new_value).first()
            model = VHOST_MODELS.get(service_domain.source.kind)
            if model is not None:
                model.objects.filter(
                    source=service_domain.source, name=service_domain.vhost_name
                ).update(service=service)
            service_domain.service_name = new_value
            service_domain.save(update_fields=["service_name"])
            changed_asset_ids.add(service_domain.source.asset_id)
            applied += 1

    return applied, len(changed_asset_ids)
