from dataclasses import dataclass

import openpyxl
from django.db.models import Count, Prefetch
from django.http import HttpResponse

from facts.dynamic_fields import coerce_fact_value
from systems.models import SystemHost, SystemHostFieldDefinition, SystemHostFieldValue, SystemSource

# SystemHost는 Asset.hostname 같은 단일 고유 식별자가 없다(external_id는 유일하지만
# vCenter/Nutanix 내부 식별자라 사람이 엑셀에 옮겨 적기엔 부적합) - source_name+name 조합을
# 매칭 키로 쓴다(서비스 조회의 (hostname, vhost) 복합 키와 같은 이유: 한 컬럼만으론 유일성이
# 보장 안 될 때 실제로 사람이 알아볼 수 있는 값끼리 묶어서 키로 삼는다).
SOURCE_NAME_HEADER = "source_name"
NAME_HEADER = "name"
MAX_ROWS = 2000

# 종류(kind)/VM 수는 FactFieldDefinition이 아니라 SystemSource/vms 관계로만 나오는 참고용
# 컬럼이라 동적 필드 목록엔 안 걸린다. AUTO 필드와 같은 취지로 다운로드엔 싣되 업로드 시엔
# 항상 무시(VM 수는 애초에 그때그때 세는 값이라 편집 대상 자체가 아님).
FIXED_REFERENCE_LABELS = ["종류", "VM 수"]


class ImportFileError(Exception):
    """헤더가 잘못됐거나 컬럼을 매칭할 수 없는 등, 파일 자체를 거부해야 하는 경우."""


@dataclass
class PendingUpdate:
    row_number: int
    host_id: int
    host_label: str
    field_id: int
    field_label: str
    old_value: str
    new_value: str


@dataclass
class InvalidCell:
    row_number: int
    host_label: str
    field_label: str
    raw_value: str


@dataclass
class ImportResult:
    updates: list[PendingUpdate]
    unmatched_hosts: list[tuple[int, str, str]]
    invalid_cells: list[InvalidCell]


def _dynamic_fields() -> list[SystemHostFieldDefinition]:
    """다운로드에는 AUTO+MANUAL 둘 다 참고용으로 실어서 대시보드 화면과 같은 값을 보여준다.
    반영 대상은 기본적으로 MANUAL뿐이지만, kind=physical(수기 등록) 호스트는 AUTO도 반영
    대상이다(아래 parse_system_host_workbook) - facts 앱의 dashboard/excel_import.py와
    비슷한 패턴이되, push 자체가 없는 kind라 AUTO를 잠글 이유가 없다는 점이 다름."""
    return list(
        SystemHostFieldDefinition.objects.filter(is_visible=True).order_by("sort_order", "id")
    )


def _dynamic_fields_by_label(
    dynamic_fields: list[SystemHostFieldDefinition],
) -> dict[str, SystemHostFieldDefinition]:
    labels = [fd.label for fd in dynamic_fields]
    duplicate_labels = sorted({label for label in labels if labels.count(label) > 1})
    if duplicate_labels:
        raise ImportFileError(
            "다음 라벨을 쓰는 필드가 여러 개 등록돼 있어 엑셀 헤더와 매칭할 수 없습니다: "
            f"{', '.join(duplicate_labels)}. admin에서 라벨을 정리한 뒤 다시 시도해주세요."
        )
    return {fd.label: fd for fd in dynamic_fields}


def _current_stored_value(host, field_definition: SystemHostFieldDefinition) -> SystemHostFieldValue | None:
    if host is None:
        return None
    return next(
        (v for v in host.field_values.all() if v.field_definition_id == field_definition.id), None
    )


def _current_value_display(host, field_definition: SystemHostFieldDefinition) -> str:
    value = _current_stored_value(host, field_definition)
    if value is None:
        return ""
    for candidate in (value.value_text, value.value_number, value.value_date):
        if candidate not in (None, ""):
            return str(candidate)
    return ""


def export_system_host_workbook() -> HttpResponse:
    """전체 물리 장비의 source_name+name(매칭 키) + 동적 필드(AUTO+MANUAL) 현재 값을 xlsx로
    내려준다. vCenter/Nutanix 행의 AUTO 필드는 참고용으로 같이 싣지만 편집 대상이 아니라서
    값을 고쳐서 올려도 parse_system_host_workbook이 조용히 무시한다 - kind=physical(수기
    등록) 행은 예외로 AUTO도 반영된다."""
    dynamic_fields = _dynamic_fields()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "시스템"
    sheet.append(
        [SOURCE_NAME_HEADER, NAME_HEADER, *FIXED_REFERENCE_LABELS, *[fd.label for fd in dynamic_fields]]
    )

    # extra/source__raw_response(JSONField)는 목록에서 안 쓰므로 defer - dashboard/queries.py의
    # get_system_host_queryset과 동일 이유(Oracle에서 NCLOB이 annotate(Count(distinct=True))의
    # GROUP BY에 걸리면 ORA-00932).
    hosts = (
        SystemHost.objects.select_related("source")
        .defer("extra", "source__raw_response")
        .annotate(vm_count=Count("vms", distinct=True))
        .prefetch_related(
            Prefetch(
                "field_values", queryset=SystemHostFieldValue.objects.select_related("field_definition")
            )
        )
        .order_by("source__name", "name")
    )
    for host in hosts:
        sheet.append([
            host.source.name,
            host.name or host.external_id,
            host.source.get_kind_display(),
            host.vm_count,
            *[_current_value_display(host, fd) for fd in dynamic_fields],
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cmdb_systems.xlsx"'
    workbook.save(response)
    return response


def parse_system_host_workbook(uploaded_file) -> ImportResult:
    try:
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True)
    except Exception as exc:  # openpyxl은 손상 파일에 다양한 예외를 던짐
        raise ImportFileError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc

    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)

    header = next(rows_iter, None)
    if (
        not header
        or len(header) < 2
        or str(header[0] or "").strip() != SOURCE_NAME_HEADER
        or str(header[1] or "").strip() != NAME_HEADER
    ):
        raise ImportFileError(
            f"첫 번째/두 번째 컬럼 헤더는 '{SOURCE_NAME_HEADER}', '{NAME_HEADER}'이어야 합니다."
        )

    fields_by_label = _dynamic_fields_by_label(_dynamic_fields())

    column_fields: list[SystemHostFieldDefinition | None] = []
    unknown_headers = []
    for col_header in header[2:]:
        col_header = str(col_header).strip() if col_header is not None else ""
        if not col_header or col_header in FIXED_REFERENCE_LABELS:
            column_fields.append(None)
            continue
        field_definition = fields_by_label.get(col_header)
        if field_definition is None:
            unknown_headers.append(col_header)
        column_fields.append(field_definition)

    if unknown_headers:
        raise ImportFileError(
            "다음 컬럼은 등록된 필드와 매칭되지 않습니다: " + ", ".join(unknown_headers)
        )

    choice_values_by_field_id = {
        fd.id: set(fd.choices.values_list("value", flat=True))
        for fd in column_fields
        if fd is not None and fd.value_type == SystemHostFieldDefinition.ValueType.CHOICE
    }

    raw_rows = []
    keys_seen = set()
    for row_number, row in enumerate(rows_iter, start=2):
        if row is None or all(cell in (None, "") for cell in row):
            continue
        source_name = str(row[0]).strip() if row[0] is not None else ""
        name = str(row[1]).strip() if row[1] is not None else ""
        if not source_name or not name:
            continue
        keys_seen.add((source_name, name))
        raw_rows.append((row_number, source_name, name, row[2:]))

        if len(raw_rows) > MAX_ROWS:
            raise ImportFileError(f"한 번에 업로드 가능한 최대 행 수({MAX_ROWS}건)를 초과했습니다.")

    hosts_by_key = {
        (host.source.name, host.name): host
        for host in SystemHost.objects.filter(
            source__name__in={k[0] for k in keys_seen}, name__in={k[1] for k in keys_seen}
        )
        .select_related("source")
        .prefetch_related(
            Prefetch(
                "field_values", queryset=SystemHostFieldValue.objects.select_related("field_definition")
            )
        )
    }

    updates: list[PendingUpdate] = []
    unmatched_hosts: list[tuple[int, str, str]] = []
    invalid_cells: list[InvalidCell] = []

    for row_number, source_name, name, values in raw_rows:
        host = hosts_by_key.get((source_name, name))
        if host is None:
            unmatched_hosts.append((row_number, source_name, name))
            continue

        host_label = f"{source_name} / {name}"
        # kind=physical은 push 자체가 없어 sync_host_fields()가 절대 안 돌기 때문에, AUTO
        # 필드도 다음 push에 덮어써질 위험이 없다 - 셀 클릭 편집(dashboard/queries.py의
        # build_system_host_rows)과 같은 예외를 여기서도 적용한다.
        is_physical = host.source.kind == SystemSource.Kind.PHYSICAL

        for field_definition, cell_value in zip(column_fields, values):
            if (
                field_definition is None
                or cell_value in (None, "")
                or (
                    field_definition.source != SystemHostFieldDefinition.Source.MANUAL
                    and not is_physical
                )
            ):
                # vCenter/Nutanix의 AUTO 필드는 참고용으로만 실려있는 컬럼이라 값이 있어도
                # 반영하지 않는다 - 어차피 다음 push 때 조용히 덮어써지므로.
                continue

            defaults = coerce_fact_value(cell_value, field_definition.value_type)
            parse_failed = (
                field_definition.value_type
                in (SystemHostFieldDefinition.ValueType.NUMBER, SystemHostFieldDefinition.ValueType.DATE)
                and all(v is None for v in defaults.values())
            )
            invalid_choice = field_definition.value_type == SystemHostFieldDefinition.ValueType.CHOICE and (
                str(cell_value) not in choice_values_by_field_id.get(field_definition.id, set())
            )
            if parse_failed or invalid_choice:
                invalid_cells.append(
                    InvalidCell(
                        row_number=row_number,
                        host_label=host_label,
                        field_label=field_definition.label,
                        raw_value=str(cell_value),
                    )
                )
                continue

            stored_value = _current_stored_value(host, field_definition)
            unchanged = stored_value is not None and all(
                getattr(stored_value, key) == value for key, value in defaults.items()
            )
            if unchanged:
                continue

            updates.append(
                PendingUpdate(
                    row_number=row_number,
                    host_id=host.id,
                    host_label=host_label,
                    field_id=field_definition.id,
                    field_label=field_definition.label,
                    old_value=_current_value_display(host, field_definition),
                    new_value=str(cell_value),
                )
            )

    return ImportResult(updates=updates, unmatched_hosts=unmatched_hosts, invalid_cells=invalid_cells)


def apply_updates(payload: list[dict]) -> tuple[int, int]:
    """미리보기에서 확정된 항목을 실제로 반영. (반영된 값 개수, 영향받은 호스트 수)를 반환.

    parse_system_host_workbook과 같은 예외(kind=physical은 AUTO도 반영)를 여기서도 다시
    검사한다 - payload는 미리보기 화면을 거쳐 브라우저에서 그대로 돌아오는 값이라, parse
    단계의 필터링만 믿지 않고 반영 시점에도 한 번 더 확인해야 vCenter/Nutanix 호스트의 AUTO
    값이 실수로라도 덮어써지지 않는다."""
    field_definitions = {
        fd.id: fd for fd in SystemHostFieldDefinition.objects.filter(is_visible=True)
    }
    host_ids = {item["host_id"] for item in payload}
    hosts = {
        host.id: host
        for host in SystemHost.objects.filter(id__in=host_ids).select_related("source")
    }

    changed_host_ids = set()
    applied = 0
    for item in payload:
        field_definition = field_definitions.get(item["field_id"])
        host = hosts.get(item["host_id"])
        if field_definition is None or host is None:
            continue
        is_physical = host.source.kind == SystemSource.Kind.PHYSICAL
        if field_definition.source != SystemHostFieldDefinition.Source.MANUAL and not is_physical:
            continue

        defaults = coerce_fact_value(item["new_value"], field_definition.value_type)
        SystemHostFieldValue.objects.update_or_create(
            host=host, field_definition=field_definition, defaults=defaults
        )
        changed_host_ids.add(host.id)
        applied += 1

    return applied, len(changed_host_ids)
