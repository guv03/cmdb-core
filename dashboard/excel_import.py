from dataclasses import dataclass

import openpyxl
from django.db.models import Prefetch
from django.http import HttpResponse
from django.utils import timezone

from core.models import Asset
from core.reconciliation import normalize_hostname
from facts.dynamic_fields import coerce_fact_value
from facts.models import FactFieldDefinition, HostFactValue

HOSTNAME_HEADER = "hostname"
MAX_ROWS = 2000

# hostname처럼 첫 컬럼은 아니지만, FactFieldDefinition이 아니라 LEADING_FIXED_COLUMNS(대시보드
# 전용 하드코딩 컬럼)라 동적 필드 목록엔 안 걸리는 참고용 컬럼. AUTO 필드와 같은 취지로 다운로드엔
# 싣되 업로드 시엔 항상 무시(편집 대상 아님, 값 자체가 push/재계산으로만 바뀜).
FIXED_REFERENCE_LABELS = ["IP", "OS"]


class ImportFileError(Exception):
    """헤더가 잘못됐거나 컬럼을 매칭할 수 없는 등, 파일 자체를 거부해야 하는 경우."""


@dataclass
class PendingUpdate:
    row_number: int
    asset_id: int
    hostname: str
    field_id: int
    field_label: str
    old_value: str
    new_value: str


@dataclass
class InvalidCell:
    row_number: int
    hostname: str
    field_label: str
    raw_value: str


@dataclass
class ImportResult:
    updates: list[PendingUpdate]
    unmatched_hostnames: list[tuple[int, str]]
    invalid_cells: list[InvalidCell]


def _dynamic_fields() -> list[FactFieldDefinition]:
    """FIXED를 뺀 전체 동적 필드(AUTO+MANUAL) - 다운로드에는 둘 다 참고용으로 실어서
    대시보드 화면과 같은 값을 보여주고, 실제 반영 대상은 이 중 MANUAL만이다(아래
    parse_manual_field_workbook)."""
    return list(
        FactFieldDefinition.objects.exclude(source=FactFieldDefinition.Source.FIXED)
        .filter(is_visible=True)
        .order_by("sort_order", "id")
    )


def _dynamic_fields_by_label(dynamic_fields: list[FactFieldDefinition]) -> dict[str, FactFieldDefinition]:
    """헤더 매칭용 라벨 사전 - AUTO/MANUAL을 합쳐서 한 시트에 같이 싣기 때문에, 라벨 중복은
    소스 구분 없이 전체를 대상으로 검사해야 한다(label은 DB 레벨 unique가 아니라서 admin에서
    실수로 겹치게 등록할 수 있음)."""
    labels = [fd.label for fd in dynamic_fields]
    duplicate_labels = sorted({label for label in labels if labels.count(label) > 1})
    if duplicate_labels:
        raise ImportFileError(
            "다음 라벨을 쓰는 필드가 여러 개 등록돼 있어 엑셀 헤더와 매칭할 수 없습니다: "
            f"{', '.join(duplicate_labels)}. admin에서 라벨을 정리한 뒤 다시 시도해주세요."
        )
    return {fd.label: fd for fd in dynamic_fields}


def _current_stored_value(host_fact, field_definition: FactFieldDefinition) -> HostFactValue | None:
    if host_fact is None:
        return None
    return next(
        (v for v in host_fact.values.all() if v.field_definition_id == field_definition.id), None
    )


def _current_value_display(host_fact, field_definition: FactFieldDefinition) -> str:
    value = _current_stored_value(host_fact, field_definition)
    if value is None:
        return ""
    for candidate in (value.value_text, value.value_number, value.value_date):
        if candidate not in (None, ""):
            return str(candidate)
    return ""


def export_manual_field_workbook() -> HttpResponse:
    """전체 자산의 hostname + 동적 필드(AUTO+MANUAL) 현재 값을 xlsx로 내려준다(서비스 조회의
    엑셀 다운로드와 동일한 취지 - 몇 칸만 고쳐서 재업로드하는 흐름). AUTO 필드는 대시보드
    화면과 맥락을 맞춰 참고용으로 같이 싣지만 편집 대상이 아니라서, 값을 고쳐서 올려도
    parse_manual_field_workbook이 조용히 무시한다(반영은 MANUAL 필드만). 헤더가 그대로
    parse_manual_field_workbook이 기대하는 업로드 형식이라 다운로드/업로드가 짝을 이룬다."""
    dynamic_fields = _dynamic_fields()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "자산"
    sheet.append([HOSTNAME_HEADER, *FIXED_REFERENCE_LABELS, *[fd.label for fd in dynamic_fields]])

    assets = (
        Asset.objects.select_related("hostfact")
        .prefetch_related(
            Prefetch("hostfact__values", queryset=HostFactValue.objects.select_related("field_definition"))
        )
        .order_by("hostname")
    )
    for asset in assets:
        host_fact = getattr(asset, "hostfact", None)
        sheet.append([
            asset.hostname,
            asset.primary_ip or "",
            getattr(host_fact, "os_family", "") or "",
            *[_current_value_display(host_fact, fd) for fd in dynamic_fields],
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cmdb_assets.xlsx"'
    workbook.save(response)
    return response


def parse_manual_field_workbook(uploaded_file) -> ImportResult:
    try:
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True)
    except Exception as exc:  # openpyxl은 손상 파일에 다양한 예외를 던짐
        raise ImportFileError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc

    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)

    header = next(rows_iter, None)
    if not header or str(header[0] or "").strip() != HOSTNAME_HEADER:
        raise ImportFileError(f"첫 번째 컬럼 헤더는 '{HOSTNAME_HEADER}'이어야 합니다.")

    # AUTO 필드 헤더도 알려진 컬럼으로 인정한다(다운로드가 AUTO도 참고용으로 같이 싣기
    # 때문에, 안 고치고 그대로 재업로드해도 "인식 못하는 컬럼"으로 걸리면 안 됨) - 실제
    # 반영 여부는 아래 행 처리 루프에서 source로 다시 가른다.
    fields_by_label = _dynamic_fields_by_label(_dynamic_fields())

    column_fields: list[FactFieldDefinition | None] = []
    unknown_headers = []
    for col_header in header[1:]:
        col_header = str(col_header).strip() if col_header is not None else ""
        if not col_header or col_header in FIXED_REFERENCE_LABELS:
            column_fields.append(None)
            continue
        field_definition = fields_by_label.get(col_header)
        if field_definition is None:
            unknown_headers.append(col_header)
        column_fields.append(field_definition)

    if unknown_headers:
        raise ImportFileError(
            "다음 컬럼은 등록된 필드와 매칭되지 않습니다: " + ", ".join(unknown_headers)
        )

    choice_values_by_field_id = {
        fd.id: set(fd.choices.values_list("value", flat=True))
        for fd in column_fields
        if fd is not None and fd.value_type == FactFieldDefinition.ValueType.CHOICE
    }

    raw_rows = []
    hostnames = set()
    for row_number, row in enumerate(rows_iter, start=2):
        if row is None or all(cell in (None, "") for cell in row):
            continue
        hostname = normalize_hostname(str(row[0])) if row[0] is not None else ""
        if not hostname:
            continue
        hostnames.add(hostname)
        raw_rows.append((row_number, hostname, row[1:]))

        if len(raw_rows) > MAX_ROWS:
            raise ImportFileError(f"한 번에 업로드 가능한 최대 행 수({MAX_ROWS}건)를 초과했습니다.")

    assets_by_hostname = {
        asset.hostname: asset
        for asset in Asset.objects.filter(hostname__in=hostnames)
        .select_related("hostfact")
        .prefetch_related(
            Prefetch(
                "hostfact__values", queryset=HostFactValue.objects.select_related("field_definition")
            )
        )
    }

    updates: list[PendingUpdate] = []
    unmatched_hostnames: list[tuple[int, str]] = []
    invalid_cells: list[InvalidCell] = []

    for row_number, hostname, values in raw_rows:
        asset = assets_by_hostname.get(hostname)
        if asset is None:
            unmatched_hostnames.append((row_number, hostname))
            continue

        host_fact = getattr(asset, "hostfact", None)

        for field_definition, cell_value in zip(column_fields, values):
            if (
                field_definition is None
                or field_definition.source != FactFieldDefinition.Source.MANUAL
                or cell_value in (None, "")
            ):
                # AUTO 필드는 참고용으로만 실려있는 컬럼이라 값이 있어도 반영하지 않는다 -
                # 어차피 다음 push 때 조용히 덮어써지므로(수기 입력과 반대로 push가 원본).
                continue

            defaults = coerce_fact_value(cell_value, field_definition.value_type)
            parse_failed = (
                field_definition.value_type
                in (FactFieldDefinition.ValueType.NUMBER, FactFieldDefinition.ValueType.DATE)
                and all(v is None for v in defaults.values())
            )
            invalid_choice = field_definition.value_type == FactFieldDefinition.ValueType.CHOICE and (
                str(cell_value) not in choice_values_by_field_id.get(field_definition.id, set())
            )
            if parse_failed or invalid_choice:
                invalid_cells.append(
                    InvalidCell(
                        row_number=row_number,
                        hostname=hostname,
                        field_label=field_definition.label,
                        raw_value=str(cell_value),
                    )
                )
                continue

            stored_value = _current_stored_value(host_fact, field_definition)
            unchanged = stored_value is not None and all(
                getattr(stored_value, key) == value for key, value in defaults.items()
            )
            if unchanged:
                continue

            updates.append(
                PendingUpdate(
                    row_number=row_number,
                    asset_id=asset.id,
                    hostname=hostname,
                    field_id=field_definition.id,
                    field_label=field_definition.label,
                    old_value=_current_value_display(host_fact, field_definition),
                    new_value=str(cell_value),
                )
            )

    return ImportResult(
        updates=updates, unmatched_hostnames=unmatched_hostnames, invalid_cells=invalid_cells
    )


def apply_updates(payload: list[dict]) -> tuple[int, int]:
    """미리보기에서 확정된 항목을 실제로 반영. (반영된 값 개수, 영향받은 자산 수)를 반환."""
    field_definitions = {
        fd.id: fd
        for fd in FactFieldDefinition.objects.filter(
            source=FactFieldDefinition.Source.MANUAL, is_visible=True
        )
    }
    asset_ids = {item["asset_id"] for item in payload}
    assets = {
        asset.id: asset
        for asset in Asset.objects.filter(id__in=asset_ids).select_related("hostfact")
    }

    changed_asset_ids = set()
    applied = 0
    for item in payload:
        field_definition = field_definitions.get(item["field_id"])
        asset = assets.get(item["asset_id"])
        if field_definition is None or asset is None:
            continue
        host_fact = getattr(asset, "hostfact", None)
        if host_fact is None:
            continue

        defaults = coerce_fact_value(item["new_value"], field_definition.value_type)
        HostFactValue.objects.update_or_create(
            host_fact=host_fact, field_definition=field_definition, defaults=defaults
        )
        changed_asset_ids.add(asset.id)
        applied += 1

    if changed_asset_ids:
        Asset.objects.filter(id__in=changed_asset_ids).update(last_changed_at=timezone.now())

    return applied, len(changed_asset_ids)
